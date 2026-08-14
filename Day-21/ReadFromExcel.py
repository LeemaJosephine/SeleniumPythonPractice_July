from openpyxl import load_workbook

# workbook = load_workbook("Login_Data.xlsx")
# sheet = workbook["Login_Data"]

# Reading Cell by Cell
# username = sheet["A2"].value
# password = sheet["B2"].value
# result = sheet["C2"].value

# Read row by row

# for column in sheet.iter_rows(min_row=2, values_only=True):
#
#     username = column[0]
#     password = column[1]
#     print(username, password)

# Reusable excel reader

def read_excel(file_path, sheet_name):

    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    data = []  # Create empty list

    for row in sheet.iter_rows(min_row=2, values_only=True): # To skip header row starting from 2
        data.append(row)

    return data

data = read_excel("TestData/Login_Data.xlsx", "Login_Data")
print(data)




