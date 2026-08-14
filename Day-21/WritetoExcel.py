from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

# workbook = Workbook()
#
# sheet = workbook.create_sheet('Login_Data')

# sheet['A1'] = "Username"
# sheet['B1'] = "Password"
# sheet['C1'] = " Status"
#
# sheet['A2'] ="standard_user"
# sheet['B2'] = "secret_password"
# sheet['C2'] = "PASS"

# Add data to existing excel

file_path = "TestData/Login_Data.xlsx"
workbook = load_workbook(file_path)
sheet = workbook["Login_Data"]  # get into the Login_Data sheet so passing the sheet name

sheet.append(["problem_user", "secret_sauce", "PASS"])

workbook.save(file_path)

