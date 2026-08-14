from openpyxl import load_workbook


def get_data_from_excel(file_path, sheet_name):

    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    rows = sheet.max_row   # No.of rows
    columns = sheet.max_column  # No.of columns
    data = []  # empty list store the values

    for row in range(2, rows + 1):  # Skipping header row
        row_data = []
        for col in range(1, columns + 1):
            value = sheet.cell(row=row, column=col).value
            if value is None:
                value = ""  # Explicitly handling nulls as empty strings
            row_data.append(value)
        data.append(tuple(row_data))

    return data

data = get_data_from_excel("TestData/Login_Data.xlsx", "Login_Data")
print(data)



